from openpyxl import load_workbook
from pdf2image import convert_from_path
from PIL import Image
from pytesseract import image_to_string
from pathlib import Path
import utils
import tempfile

try:
    from file_processing_utils import TEMP_IMAGE_DIR
except Exception:
    TEMP_IMAGE_DIR = Path(tempfile.gettempdir()) / "tempimage"
    TEMP_IMAGE_DIR.mkdir(parents=True, exist_ok=True)

def extract_text_from_excel(file_path):
    """
    Extrahiert Text aus Excel-Dateien.
    """
    try:
        wb = load_workbook(filename=file_path, data_only=True)
        extracted_text = ""
        for sheet in wb.sheetnames:
            extracted_text += f"\n--- Sheet: {sheet} ---\n"
            for row in wb[sheet].iter_rows(values_only=True):
                extracted_text += "\t".join(str(cell) if cell else "" for cell in row) + "\n"
        return extracted_text
    except Exception as e:
        utils.log_error(f"Fehler beim Lesen von Excel-Datei {file_path}: {e}")
        return ""

def extract_text_from_markdown(file_path):
    """
    Extrahiert Text aus Markdown-Dateien.
    """
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    except Exception as e:
        utils.log_error(f"Fehler beim Lesen von Markdown-Datei {file_path}: {e}")
        return ""

def extract_text_from_pdf_pages(file_path, poppler_path=None):
    """
    Extrahiert Text aus PDF-Seiten mithilfe von OCR.
    Jedes Bild wird als PNG im TEMP_IMAGE_DIR gespeichert und anschließend per Tesseract gelesen.
    """
    try:
        pages = convert_from_path(file_path, poppler_path=poppler_path)
        extracted_text = ""
        for page_num, page in enumerate(pages):
            temp_image_path = TEMP_IMAGE_DIR / f"{Path(file_path).stem}_page_{page_num:03}.png"
            page.save(temp_image_path, "PNG")
            with Image.open(temp_image_path) as img:
                extracted_text += image_to_string(img, lang="deu+eng") + "\n"
        return extracted_text
    except Exception as e:
        utils.log_error(f"Fehler beim Verarbeiten der PDF {file_path}: {e}")
        return ""
